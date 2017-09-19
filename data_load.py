# -*- coding: utf-8 -*-
"""
Created on Mon Jul 10 09:27:59 2017

@author: root
"""
import openpyxl 
import warnings
import numpy as np
import os 
#%%
path2data = os.path.split(os.path.abspath("__file__"))[0]+r"\AD"
_wb = openpyxl.load_workbook(path2data+r"\Kraftwerksparameter_2016.xlsx")
_kwk_param = _wb.get_sheet_by_name(_wb.get_sheet_names()[0])
_var = {_kwk_param .cell(row=2,column=i).value:i for i in range(2,11)}
#dic[newkey] = dic.pop(oldkey)  ##
tec = [_kwk_param .cell(row=i,column=1).value for i in range(3,14)]

Pl= {varname: [_kwk_param .cell(row=i,column=col).value for i in range(3,14)] for varname,col in _var.items()}

Pl_dict = {varname:dict(zip(tec, liste)) for varname,liste in Pl.items()}    

_var2 = {_kwk_param .cell(row=17,column=i).value:i for i in range(2,10)}
Ws = {name:_kwk_param .cell(row=18,column=col).value for name,col in _var2.items() }
#%%
_wb = openpyxl.load_workbook(path2data+r"\Szenarien_2016.xlsx")
_tabelle = _wb.get_sheet_by_name(_wb.get_sheet_names()[0])

#Jahr = [_tabelle.cell(row=1,column=col).value for col in range(4,9)]
Jahr = [_tabelle.cell(row=15,column=col).value for col in range(3,10)]
szenario = ['Base','M2M','Eff']

def Sz(z): 
    x= [_tabelle.cell(row=z+i*4,column=col).value for i in range(0,3) for col in range(3,10) ] 
    y= {szenario[i]:dict(zip(Jahr,[ _tabelle.cell(row=z+i*4,column=col).value  for col in range(3,10)])) for i in range(0,3) }
    z = {(szenario[i],Jahr[col-3]): _tabelle.cell(row=z+i*4,column=col).value  for col in range(3,10) for i in range(0,3) } 
    return (x,y,z)

S = {"Name": szenario,
     "P_oil":Sz(16)[0],
     "P_gas":Sz(17)[0],
     "P_coal":Sz(18)[0],
     "P_co2":Sz(19)[0],
     }

S_dict = {"Name": szenario,
     "P_oil":Sz(16)[2],
     "P_gas":Sz(17)[2],
     "P_coal":Sz(18)[2],
     "P_co2":Sz(19)[2],
     }

em = {"gas":0.2,
      "coal":0.4,
      "oil":0.3,
      "muell":0.2,
      }
#%%
warnings.simplefilter("ignore")   
# Merkwürdige Excel File - braucht lange zum laden und gibt ein Warning aus ..keine Ahnung warum ? 
S["strompreis"] = {}
S_dict["strompreis"] = {}
_wb = openpyxl.load_workbook(path2data+r"\Preise.xlsx")

for j in range(0,4):
    _tabelle = _wb.get_sheet_by_name(_wb.get_sheet_names()[j])
    
    S["strompreis"][ _wb.get_sheet_names()[j]] = [[_tabelle.cell(row=z,column=col).value for col in range(2,5)] for z in  range(2,_tabelle.max_row+1) ]
#    S_dict["strompreis"][ _wb.get_sheet_names()[j]]= {szenario[col-2]:dict(zip(range(1,87661),[_tabelle.cell(row=z,column=col).value for z in  range(2,_tabelle.max_row+1)])) for col in range(2,5)}  
    for col in range(2,5):
        for z in  range(2,_tabelle.max_row+1):
            S_dict["strompreis"][(int( _wb.get_sheet_names()[j]),szenario[col-2],z-1)] = _tabelle.cell(row=z,column=col).value 
    
#%%
_wb = openpyxl.load_workbook(path2data+r"\Waermebedarf_WIEN_in_MWh.xlsx")
_tabelle = _wb.get_sheet_by_name(_wb.get_sheet_names()[0])
demand_th = [_tabelle.cell(row=r,column=1).value for r in range(1,_tabelle.max_row+1)]
demand_th_dict = {i+1:demand_th[i] for i in range(0,len(demand_th))}
#%%
t_v = np.load(path2data+r"\time_vec.npy")
#%%
radiation = np.load(path2data+r"\Strahlung.npy")
radiation_dic = dict(zip(range(1,8761),radiation.reshape(1,8760).tolist()[0]))
#%%
#%%

Ws["IK_WS_KAP"] = Ws["IK_WS_KAP"]*0.5
Ws["IK_WS_P"] = Ws["IK_WS_P"]*0.5

#%% Add paramters
z=0.05 #% Kalkulationszinssatz

#%% Optinale Förderung von Solarthermie
Foerderung_Solar=0; #% €/kW Investitionsförderung Solarthermie 

#%% Geforderter Mindestanteil von Erneuerbaren
min_anteil_erneuerbar=0;

#%% Jährliche Kapitalkosten über Annuität in €/MWh

Pl["Kj"]=list(np.array(Pl["IK"])*((1+z)**np.array(Pl["LD"])*z/((1+z)**np.array(Pl["LD"])-1))*1000)


#% Kapitalkosten der Solarthermie nach Förderung
Pl["Kj"][2]=(Pl["IK"][2]-Foerderung_Solar)*((1+z)**Pl["LD"][2]*z/((1+z)**Pl["LD"][2]-1))*1000;

#%Kapitalkosten Wärmespeicher
Ws["Kj_kap"] = Ws["IK_WS_KAP"]*((1+z)**Ws["LD"]*z/((1+z)**Ws["LD"]-1))*1000;
Ws["Kj_P"]=Ws["IK_WS_P"]*((1+z)**Ws["LD"]*z/((1+z)**Ws["LD"]-1))*1000;

Ws["Entladeleistung_bestand"]=140;
#%%
#%% Definition der Grenzkosten
'''
grenzkosten werden für alle Abwärmequellen und Geothermie jeweils nur einmal berechnet
da für alle Potenzialstufen die gleichen Wirkungsgrade
angenommen wurden - die Wirkungsgrade sind hier als Aufwand für 
Pumpen usw zu verstehen und sind so wir der COP einer Wärmepumpe
zu interpretieren
'''
#%AUSWAHL 
year = 2020
scenario = "M2M"
mc = {}
for j in [tec[0]]:
    if j in ["Spitzenlastkessel"]:
        for t in range(1,8760+1):
            mc[j,t]= S_dict["P_gas"][scenario,year] /  \
                          Pl_dict["n_th"][j] + \
                          em["gas"]*S_dict["P_co2"][scenario,year] / \
                          Pl_dict["n_th"][j]   
    elif j in ["Müllverbrennung"]:
        for t in range(1,8760+1):
            mc[j,t] =5
    else:
        for t in range(1,8760+1):   
            mc[j,t] = S_dict["strompreis"][year,scenario,t] /Pl_dict["n_th"][j]

#%%
#%%
'''
Definition des Zeitraums in dem Wärmepumpen nicht verfügbar sind
und Spitzenlasten im Stromsystem auftreten (in dieser Zeit ist
eine zusätzliche elektrische Last für die Deckung der
Wärmenachfrage problematisch - deshalb werden hier auch Kosten
für P_el_max angesetzt
'''
t_knappheit = list(range(1,60*24+1))+list(range(8760-24*30,8760+1))

'''
Wärmepumpe ist hier nicht verfügbar
Constraints = [Constraints, x_WP(t_knappheit) <= 0];
'''
#%%
def import_data():        
    return Pl,Pl_dict,Ws,Jahr,S,S_dict,em,demand_th,demand_th_dict,tec,szenario,t_v,radiation,radiation_dic,mc,year,scenario,min_anteil_erneuerbar,t_knappheit

